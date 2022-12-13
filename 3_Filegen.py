# These are the two libraries we will need. OpenPyxl is what you need to install, and "os" comes with Python by default.
import openpyxl
import os


def parse_date_from_file_name(file_name):
    """
    This is a simple function to automatically determine the date for a file by looking at the name of that file.
    This will only work if there are exactly 2 underscores in the name of the file, and they are on either side of the date.

    For example, a file named:
    excel_file_2022_01_10_with_stuff.xlsx
    WILL NOT WORK.

    Instead it should be named:
    excel file_20220110_with stuff.xlsx

    :param file_name: Name of the file to find the date from.
    :return: The date detected from the file's name.
    """

    print("Parsing date from file name {}...".format(file_name))
    idx1 = file_name.find("_") + 1
    idx2 = file_name.rfind("_")

    date = file_name[idx1:idx2]

    print("Date detected as: {}".format(date))
    return date


def generate_filenames_automatically(workbook):
    """
    This function will scan the matrix located in the first sheet of the excel file and map cells in the matrix to
    patient IDs and markers. This will be necessary when we want to name our new files later on. Make sure every Excel
    file you want to run this code on contains a matrix in the same format located in Sheet1!

    :param workbook: The Excel file we are copying data from.
    :return: A Python dictionary which maps cells to patient IDs and marker names.
    """

    matrix_column_range = "A:N"
    matrix_sheet_name = "Sheet1"

    matrix_sheet = workbook[matrix_sheet_name]
    matrix_cell_to_patient_id_map = {}
    matrix_letters = []
    matrix_numbers = []
    matrix_patient_ids = []
    matrix_marker_names = []

    print("Checking columns {} in sheet {} for matrix...".format(matrix_column_range, matrix_sheet_name))

    # Loop over columns A and B in the worksheet.
    for letter_cell, patient_id_cell in zip(matrix_sheet["A"], matrix_sheet["B"]):

        # Check each row in columns A and B to find the letter and patient ID associated with that row in the matrix.
        letter = letter_cell.value
        patient_id = patient_id_cell.value

        # If there is an entry in each row, we'll add the letter and patient ID to our lists for later.
        if letter is not None and patient_id is not None:
            matrix_letters.append(letter)
            matrix_patient_ids.append(patient_id)

    # Loop over each column in the matrix.
    for col in matrix_sheet[matrix_column_range]:

        # Check rows 1 and 2 at each column to find the number and marker name associated with that column in the matrix.
        number = col[0].value
        marker_name = col[1].value

        # If the matrix has an entry here, we'll add the number and marker name to our lists for later.
        if number is not None and marker_name is not None:
            matrix_numbers.append(number)
            matrix_marker_names.append(marker_name)

    # Now that we have found the patient IDs, marker names, and their associated cell names in the matrix, we will
    # loop over them all and create a map that will let us find the patient ID and marker name for any cell.
    # When we're done we can ask the map for the ID and marker associated with "A1" (or any other cell name)
    # and it will tell us  "Li13" is the patient ID and "RBP4" is the marker name.
    for letter, patient_id in zip(matrix_letters, matrix_patient_ids):

        # Loop over each number and marker name.
        for number, marker_name in zip(matrix_numbers, matrix_marker_names):

            # Create the name of each cell by combining the letter of the row with the number of the column.
            matrix_cell_name = "{}{}".format(letter, number)

            # Create the patient ID and marker name associated with each cell.
            patient_id_and_data_type = [patient_id, marker_name]

            # Store the cell name, patient ID, and marker name in our map.
            matrix_cell_to_patient_id_map[matrix_cell_name] = patient_id_and_data_type


    for key, value in matrix_cell_to_patient_id_map.items():
        print("Cell name {} mapped to patient ID {} and marker {}".format(key, value[0], value[1]))

    return matrix_cell_to_patient_id_map


def read_matrix_for_filenames(workbook):
    """
    This is the SECOND METHOD for matching patient IDs with marker types from the matrix in Sheet1 of the Excel file being
    read. It expects that the matrix will be in Sheet1, and span columns A to N. This code expects that there are letters
    in row 1, there is NOTHING in row 2, and the matrix begins on row 3. Column 1 should contain letters, and the matrix
    should begin on column 2.

    :param workbook:
    :return:
    """

    matrix_column_range = "A:N"
    matrix_sheet_name = "Sheet1"

    matrix_sheet = workbook[matrix_sheet_name]
    letters = matrix_sheet["A"][2:]
    matrix_cell_to_patient_id_map = {}

    print("Checking columns {} in sheet {} for matrix...".format(matrix_column_range, matrix_sheet_name))

    # Loop over columns A:N.
    for col in matrix_sheet[matrix_column_range]:

        # Trim each column to begin where the matrix does. This expects the matrix to start on row 3.
        trimmed_column = col[2:]

        # Read the number from row 1 in each column.
        number = col[0].value

        # If there is no number here, we'll skip this column and move on. This code will allow us to skip columns 1 and 2.
        if number is None:
            continue


        # Scan down the trimmed column.
        for i in range(len(trimmed_column)):

            # Read the letter in column 1 that matches this row in the matrix.
            letter = letters[i].value

            # If there is no letter in column 1 at this row, it means we've reached the end of the matrix along this
            # column, so we will stop scanning down this column and move on to the next one.
            if letter is None:
                break

            # If there is no entry in the cell we're currently looking at, just ignore it and move to the next cell in this column.
            if trimmed_column[i].value is None:
                continue

            # Read the patient ID and marker type from each cell in the matrix.
            patient_id_and_data_type = list(trimmed_column[i].value.split("_"))

            # Create the name of each cell by combining the letter of the row with the number of the column.
            matrix_cell_name = "{}{}".format(letter, number)

            # Store the cell name, patient ID, and marker name in our map.
            matrix_cell_to_patient_id_map[matrix_cell_name] = patient_id_and_data_type

    for key, value in matrix_cell_to_patient_id_map.items():
        print("Cell name {} mapped to patient ID {} and marker {}".format(key, value[0], value[1]))

    return matrix_cell_to_patient_id_map


def create_new_xlsx_file(workbook, matrix_cell_name, copy_from_worksheet, new_file_path):
    """
    This function will create a single new Excel file for us containing the data we want.

    :param workbook: The Excel file to copy data from.
    :param matrix_cell_name: The name of the cell in the matrix from Sheet1 to examine (A1 or B1 or C1 or D1 etc).
    :param copy_from_worksheet: Which worksheet in the Excel file we're copying (FITC_post or PE_post).
    :param new_file_path: File path to save our new file to. This will be automatically generated, so don't worry about it!
    :return:
    """

    # Before we start we'll just quickly check to make sure that the worksheet we want to copy data from actually exists.
    if copy_from_worksheet not in workbook.sheetnames:
        print("Could not locate worksheet named {} in Excel file!".format(copy_from_worksheet))
        return

    print("Creating new file for data at cell {} in worksheet {}. File path: {} ".format(matrix_cell_name,
                                                                                         copy_from_worksheet,
                                                                                         new_file_path))
    # Open up a new empty Excel file.
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active

    # Access the worksheet from the Excel file we want to copy from.
    ws = workbook[copy_from_worksheet]

    # Here we set up some variables we'll need when checking the data.
    col_to_fill = 1
    target_patient = "_{}_".format(matrix_cell_name)

    # Loop over every column in the worksheet we're copying from.
    for col in ws.columns:

        # The name of the column should be its first entry.
        column_name = col[0].value

        # If the column name isn't None (e.g. there is no data in that cell) and the column's name contains the patient
        # we're interested in, we'll copy data from it.
        if column_name is not None and target_patient in column_name:

            # Cut off the first row in every column (to avoid copying the name of the column to the new file).
            clipped_column = col[1:]

            # Loop over the entries in the column that contain the data we want.
            for i in range(len(clipped_column)):

                # Pull out the data from each cell.
                data = clipped_column[i].value

                # Write the data to our new file.
                new_ws.cell(row=i+1, column=col_to_fill, value=data)
            col_to_fill += 1

    # Save our new file and be done!DataD
    new_wb.save(new_file_path)


def main():
    """
    This is the function that will be run when the program starts. It will scan a folder that should contain Excel files
    like the one you sent me and automatically split them all into new Excel files in the format we discussed.

    :return:
    """

    # This is a list of the names of the worksheets we should be copying data from. In our discussion earlier you pointed
    # me toward FITC and PE, but if you have other worksheets in some files you should include them in this list. To add
    # new worksheets to this list, just add the name of each worksheet followed by a comma to the existing list. For example:
    # worksheets_to_copy = ["FITC", "PE", "some_other_worksheet", "another_worksheet", "etc"]. Be careful not to accidentally
    # include the comma inside the quotation marks!
    worksheets_to_copy = ["FITC_post", "PE_post"]

    # This is the name of the folder we will copy files from. You can put one Excel file in here at a time if you want, or
    # put every file you have in here and the code will split all of them into new files.
    folder_to_scan = "JS"

    # This is the name of the folder we will save the new Excel files to.
    save_location = "JS_folder"


    # MATRIX READING METHOD 1 (12 x 8 matrix):
    #matrix_loading_method = generate_filenames_automatically

    # MATRIX READING METHOD 2 (individual 96 well):
    matrix_loading_method = read_matrix_for_filenames

    # This code will just create the save location if it doesn't exist already.
    if not os.path.exists(save_location):
        os.makedirs(save_location)

    # Here the code is making sure that the folder you want to load Excel files from actually exists.
    assert os.path.exists(folder_to_scan), "UNABLE TO LOCATE FOLDER {}".format(folder_to_scan)

    # Scan over all the files in the folder we've given it.
    for file_name in os.listdir(folder_to_scan):
        if not (".xlsx" in file_name or ".xlsm" in file_name or ".xltx" in file_name):
            print("SKIPPING INVALID FILE:",file_name)
            continue

        # Here we find the full path for the file we want to load and open it with the OpenPyxl package.
        file_path = os.path.join(folder_to_scan, file_name)
        print("Valid file detected. Check status: {}".format(".xlsx" in file_name or ".xlsm" in file_name or ".xltx" in file_name))
        print("Loading file {}...".format(file_name))
        workbook = openpyxl.load_workbook(file_path)

        # This calls on our earlier function to automatically locate the date by examining the name of an Excel file.
        date = parse_date_from_file_name(file_name)

        # Next we'll build a map from the matrix of patient IDs and marker names to cells with the function we made earlier.
        matrix_cell_patient_id_marker_map = matrix_loading_method(workbook)

        # Pull out just the cell names from our map.
        matrix_cell_names = tuple(matrix_cell_patient_id_marker_map.keys())

        # Loop over every cell name we discovered and make an Excel file for it.
        for cell_name in matrix_cell_names:

            # Here we're looping over every worksheet we were told to examine with the `worksheets_to_copy` list.
            for worksheet_name in worksheets_to_copy:

                # First we'll find the patient ID and marker name for the cell we're looking at
                patient_id, marker = matrix_cell_patient_id_marker_map[cell_name]

                # Now we'll automatically build the name of the new file we're about to make.
                new_file_name = "[{}]_{}{}_{}.xlsx".format(date, patient_id, worksheet_name[0], marker)

                # Next we just put together the full file path for the new file.
                new_file_path = os.path.join(save_location, new_file_name)

                # Finally we call our function to copy data and create a new file for us.
                create_new_xlsx_file(workbook=workbook, matrix_cell_name=cell_name,
                                     copy_from_worksheet=worksheet_name, new_file_path=new_file_path)
        print("Finished creating new files from the file at {}!\n".format(file_path))
    print("All files successfully created!")


# This code is just the standard way to show Python where the program is supposed to begin.
if __name__ == "__main__":
    # We just call the main function and the program will run!
    main()
